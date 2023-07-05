import sys
import os
import openpyxl

from dotenv import load_dotenv, dotenv_values
from utils.query_fb import queryFb
from utils.sendmail import sendMail
from datetime import date
os.chdir('S:/Wolt_torzs')

sys.path.append('utils')

load_dotenv()
envConfig = dotenv_values(".env")

def main():

    today = date.today()
    
    """ create feldolgozott folder if not exist"""
    if not os.path.exists('feldolgozott'):
        os.mkdir('feldolgozott')

    """ get actual xls filename """
    for files in os.listdir('feldolgozando'):
        if files.endswith(".xlsx"):

            """ validate data """
            for filename in os.listdir('feldolgozando'):
            
                if filename.endswith(".xlsx"): 

                    fileNameDir= f"feldolgozando/"+filename

                    #open xlsx
                    wb = openpyxl.load_workbook(fileNameDir)

                    # select the first sheet
                    ws = wb.worksheets[0]
        
                    # iterate through row2 column3 reverese for delete rows
                    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))
                    for revrow in reversed(rows):
                        
                        validate = queryFb("one", """ SELECT CIK.ID, CIK.NEV, CIKKOD.KOD 
                                    FROM CIK
                                    JOIN CIKMNY ON CIKMNY.CIK_ID=CIK.ID
                                    JOIN CIKKOD ON CIKKOD.CIKMNY_ID=CIKMNY.ID
                                    WHERE CIKKOD.KPC_ID = 10 AND CIKKOD.KOD = '%s' """ % (revrow[2].value), envConfig)
                        print(validate)
                        if validate:
                            ws.delete_rows(revrow[0].row, 1)
                        else :
                            ws.cell(row=revrow[0].row, column=18).value = "nincs"
                    
                    # save the file
                    path = f"./feldolgozott/" + str(today) + "-" + filename
                    newFileName = str(today) + "-" + filename
                    wb.save(path)

                    # move feldolgozott
                    os.remove(fileNameDir)

                    # send as attachement
                    sendMail(path, newFileName, today, envConfig)


if __name__ == "__main__":
    main()
